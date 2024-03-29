from openpyxl import load_workbook

def colour_of_cell(cell):
    argb = cell.fill.bgColor.rgb
    print('colour: ' + argb)
    if argb[0:2] == '00':
        return None
    return argb[2:8]


class Sheet:
    def __init__(self, filename):
        self._workbook = load_workbook(filename)

    @property
    def _sheet(self):
        return self._workbook.active

    def cell_note(self, col, row):
        note = self._sheet.cell(row=row+1, column=col+1).comment
        if note:
            return str(note.content)
        return None

    def cell_colour(self, col, row):
        return colour_of_cell(self._sheet.cell(row=row+1, column=col+1))

    @property
    def columns(self):
        cols = []
        for col_index in range(self._sheet.min_column, self._sheet.max_column + 1):
            column = Column()
            cols.append(column)
            for row_index in range(self._sheet.min_row, self._sheet.max_row + 1):
                column.add_cell(self._sheet.cell(row=row_index, column=col_index))

        return cols


class Column:
    def __init__(self):
        self._cells = []

    def add_cell(self, excel_cell):
        self._cells.append(Cell(excel_cell))

    def cell(self, index):
        return self._cells[index]

    @property
    def row_count(self):
        return len(self._cells)


class Cell:
    def __init__(self, cell):
        self._cell = cell

    @property
    def content(self):
        return self._cell.value

    @property
    def colour(self):
        return colour_of_cell(self._cell)
